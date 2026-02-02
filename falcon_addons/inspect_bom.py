import openpyxl
import sys

try:
    wb = openpyxl.load_workbook('/mnt/extra-addons/Example BoM.xlsx', read_only=True, data_only=True)
    for sheet in wb.sheetnames:
        print(f"--- Sheet: {sheet} ---")
        ws = wb[sheet]
        rows = list(ws.iter_rows(max_row=5, values_only=True))
        for i, row in enumerate(rows):
            print(f"Row {i}: {row}")
except Exception as e:
    print(f"Error: {e}")
