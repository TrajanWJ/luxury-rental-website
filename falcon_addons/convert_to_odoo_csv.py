import openpyxl
import csv
import sys
import re

SOURCE_FILE = '/mnt/extra-addons/Example BoM.xlsx'
DEST_FILE = '/mnt/extra-addons/products_import_ready.csv'

print(f"Reading from {SOURCE_FILE}...")
try:
    wb = openpyxl.load_workbook(SOURCE_FILE, data_only=True)
except Exception as e:
    print(f"Failed to load source: {e}")
    sys.exit(1)

# Defined Headers for Odoo Import
HEADERS = [
    "Name",
    "Can be Purchased",
    "Product Type",
    "Tracking",
    "Cost",
    "Internal Reference",
    "Vendor / Name",
    "Vendor / Product Code",
    "Vendor / Product Name",
    "Vendor / Minimum Quantity",
    "Vendor / Delivery Lead Time",
    "Routes"
]

all_rows = []

def clean_route(val):
    if not val: return ""
    # "Arc34: Receive in 2 steps" -> "Receive in 2 steps"
    return str(val).replace("Arc34: ", "")

def map_sheet(sheet_name, header_row_idx):
    if sheet_name not in wb.sheetnames:
        print(f"Sheet {sheet_name} not found.")
        return

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    
    if len(rows) <= header_row_idx:
        return

    # Map source headers to indices
    src_headers = list(rows[header_row_idx])
    # Normalize headers for finding
    src_map = {str(h).strip(): i for i, h in enumerate(src_headers) if h}

    # Helper to get val
    def get_val(row, col_name, default=None):
        idx = src_map.get(col_name)
        if idx is not None and idx < len(row):
            return row[idx]
        return default

    # Iterate data
    for row in rows[header_row_idx+1:]:
        if not any(row): continue
        
        # Primary key: Name
        # Try "Product Name" (Sheet 1) or "Cable Name" (Sheet 2)
        name = get_val(row, "Product Name") or get_val(row, "Cable Name")
        if not name:
            continue

        item = {}
        item["Name"] = name
        
        # Purchase: Source "Purchase" (True/False) or default True ??
        # Sheet 1 has "Purchase" column. Sheet 2 does not.
        p_val = get_val(row, "Purchase")
        if p_val is True or p_val == "True":
            item["Can be Purchased"] = "TRUE"
        else:
            # If explicit False, FALSE. If missing (Sheet 2), maybe Default to TRUE? 
            # Cables usually purchased. 
            item["Can be Purchased"] = "TRUE" 

        # Product Type: "Goods" -> "Storable Product"
        pt_val = get_val(row, "Product Type")
        if pt_val == "Goods":
            item["Product Type"] = "Storable Product"
        else:
            item["Product Type"] = "Storable Product" # Default per instructions

        # Tracking: source "Track Inventory" -> "No Tracking" per example
        item["Tracking"] = "No Tracking"

        # Cost: "Cost" (Sheet 1) or "Price" (Sheet 2)
        cost = get_val(row, "Cost") or get_val(row, "Price")
        item["Cost"] = cost if cost is not None else 0.0

        # Reference
        item["Internal Reference"] = get_val(row, "Reference", "")

        # Vendor Info (Sheet 1 mainly)
        item["Vendor / Name"] = get_val(row, "Vendor", "")
        item["Vendor / Product Code"] = get_val(row, "Vendor Product Code", "")
        item["Vendor / Product Name"] = get_val(row, "Vendor Product Name", "")
        item["Vendor / Minimum Quantity"] = get_val(row, "Quantity", "") # Vendor Qty
        item["Vendor / Delivery Lead Time"] = get_val(row, "Lead Time", "")

        # Routes: Sheet 1 has 2 "Routes" columns.
        # find all indices for "Routes"
        routes_vals = []
        for i, h in enumerate(src_headers):
            if str(h).strip() == "Routes" and i < len(row):
                 val = row[i]
                 if val:
                     routes_vals.append(clean_route(val))
        
        # Unique and join
        unique_routes = []
        for r in routes_vals:
            if r not in unique_routes:
                 unique_routes.append(r)
        
        item["Routes"] = ",".join(unique_routes)

        all_rows.append([item.get(h, "") for h in HEADERS])

# Process Sheet "Three Window_Covert" (Header at row 1)
map_sheet("Three Window_Covert", 1)

# Process Sheet "Cables" (Header at row 0)
map_sheet("Cables", 0)

with open(DEST_FILE, 'w', newline='', encoding='utf-8') as f:
    writer = csv.writer(f, quoting=csv.QUOTE_ALL)
    writer.writerow(HEADERS)
    writer.writerows(all_rows)

print(f"Generated {len(all_rows)} rows in {DEST_FILE}")
