import openpyxl
from openpyxl import Workbook
import sys

# Paths
SOURCE_FILE = '/mnt/extra-addons/Example BoM.xlsx'
DEST_FILE = '/mnt/extra-addons/product_template.xlsx'

print(f"Reading from {SOURCE_FILE}...")
try:
    src_wb = openpyxl.load_workbook(SOURCE_FILE, data_only=True)
except Exception as e:
    print(f"Failed to load source: {e}")
    sys.exit(1)

tgt_wb = Workbook()
ws = tgt_wb.active
ws.title = "Produits - MP"

# Desired standard headers for Odoo Import (using accepted aliases)
STANDARD_HEADERS = ["Nom du produit", "Unité", "Prix", "Catégorie"]
# Corresponding to: name, uom_name, list_price, category_name

all_items = []
seen_headers = set(STANDARD_HEADERS)

def process_sheet(sheet_name, name_col, cost_col, header_row_idx):
    if sheet_name not in src_wb.sheetnames:
        print(f"Sheet {sheet_name} not found, skipping.")
        return

    print(f"Processing sheet: {sheet_name}...")
    src_ws = src_wb[sheet_name]
    rows = list(src_ws.iter_rows(values_only=True))
    
    if len(rows) <= header_row_idx:
        print("Empty sheet or no data.")
        return

    src_headers = rows[header_row_idx]
    # Clean headers (None to "")
    src_headers = [(h if h else f"Col_{i}") for i, h in enumerate(src_headers)]

    data_start = header_row_idx + 1
    
    count = 0
    for r in rows[data_start:]:
        # Skip if row is purely empty
        if not any(r): continue
        
        # Skip if name is empty
        try:
            name_idx = src_headers.index(name_col)
            name_val = r[name_idx]
            if not name_val:
                continue
        except ValueError:
            print(f"Column '{name_col}' not found in headers {src_headers}")
            return

        item = {}
        item["Nom du produit"] = name_val
        item["Unité"] = "Units" # Default UoM
        item["Catégorie"] = "Raw Materials" # Default Category
        
        # Price/Cost
        if cost_col in src_headers:
            idx = src_headers.index(cost_col)
            item["Prix"] = r[idx] if r[idx] is not None else 0.0
        else:
            item["Prix"] = 0.0

        # Extras
        for i, val in enumerate(r):
            h = src_headers[i]
            # Avoid overwriting standard fields, but capture original data as extra columns
            if h not in item: 
                 # If header conflicts with standard, rename it?
                 # Actually, we just add it to item. 
                 # If h is "Product Name", we mapped it to "Nom du produit".
                 # But user wants "extra fields labeled".
                 # So we keep "Product Name" as another column? PROBABLY useful for reference.
                 if h not in seen_headers:
                     seen_headers.add(h)
                 item[h] = val
        
        all_items.append(item)
        count += 1
    print(f"Extracted {count} items from {sheet_name}")

# Logic based on inspect output
# Sheet "Three Window_Covert": Headers at Row 1 (0-indexed? Inspect said Row 1).
# Inspect output: Row 0 ... Row 1: Product Name...
# So process header_row_idx=1 for Three Window_Covert
process_sheet("Three Window_Covert", "Product Name", "Cost", 1)

# Sheet "Cables": Inspect output Row 0: Cable Name...
# So header_row_idx=0
process_sheet("Cables", "Cable Name", "Price", 0)

# Write to Target
final_headers = list(STANDARD_HEADERS)
# Add extra headers found (sorted for consistency)
extras = sorted([h for h in seen_headers if h not in final_headers])
final_headers.extend(extras)

ws.append(final_headers)
for item in all_items:
    row = [item.get(h, "") for h in final_headers]
    ws.append(row)

print(f"Writing {len(all_items)} rows to {DEST_FILE}...")
tgt_wb.save(DEST_FILE)
print("Done.")
